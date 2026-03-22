import pandas as pd
import chardet
import io

from openpyxl import load_workbook


def detect_encoding(file_bytes: bytes) -> str:
    result = chardet.detect(file_bytes)
    return result.get("encoding", "utf-8") or "utf-8"


# 常见的空值标记（MySQL \N、NA、NULL 等）
_NA_VALUES = {"\\N", "\\n", "NULL", "null", "None", "none", "NA", "N/A", "n/a", "NaN", "nan", ""}


def _read_excel_with_merges(raw: bytes) -> tuple[pd.DataFrame, list[tuple]]:
    """
    用 openpyxl 读取 Excel，将合并单元格的值填充到所有被合并的位置。
    返回 (DataFrame, merge_ranges)，merge_ranges 是合并区域列表，
    每项为 (min_row, min_col, max_row, max_col)，行列从0开始（相对数据区域，不含表头）。
    """
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    # 第1步：收集所有合并区域及其左上角的值
    merge_fills = []  # [(min_row, min_col, max_row, max_col, value)]
    for merged_range in list(ws.merged_cells.ranges):
        min_r = merged_range.min_row
        min_c = merged_range.min_col
        max_r = merged_range.max_row
        max_c = merged_range.max_col
        value = ws.cell(min_r, min_c).value
        merge_fills.append((min_r, min_c, max_r, max_c, value))

    # 第2步：解除所有合并
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # 第3步：将左上角的值填充到整个合并区域
    for min_r, min_c, max_r, max_c, value in merge_fills:
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                ws.cell(r, c, value=value)

    # 第4步：读取完整数据
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    wb.close()

    if not data:
        return pd.DataFrame(), []

    # 第一行作为列名
    headers = [str(h) if h is not None else f"列{i+1}" for i, h in enumerate(data[0])]
    seen = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)

    df = pd.DataFrame(data[1:], columns=unique_headers)

    # 转换合并区域坐标：从 Excel 1-based 转为数据区域 0-based（去掉表头行）
    data_merges = []
    for min_r, min_c, max_r, max_c, _ in merge_fills:
        # 跳过纯表头行的合并
        if max_r <= 1:
            continue
        # 转为 0-based 数据行索引（Excel 第2行 = 数据第0行）
        d_min_r = max(min_r - 2, 0)
        d_max_r = max_r - 2
        d_min_c = min_c - 1
        d_max_c = max_c - 1
        if d_max_r >= 0:
            data_merges.append((d_min_r, d_min_c, d_max_r, d_max_c))

    return df, data_merges


def load_table(uploaded_file) -> tuple[pd.DataFrame, list[tuple]]:
    """
    加载表格文件。
    返回 (DataFrame, merge_ranges)，merge_ranges 为合并区域列表（可能为空）。
    """
    name = uploaded_file.name.lower()
    raw = uploaded_file.getvalue()
    merges = []

    if name.endswith(".csv"):
        encoding = detect_encoding(raw)
        df = pd.read_csv(io.BytesIO(raw), encoding=encoding, na_values=_NA_VALUES, keep_default_na=True)
    elif name.endswith((".xlsx", ".xls")):
        if name.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(raw), na_values=_NA_VALUES, keep_default_na=True)
        else:
            df, merges = _read_excel_with_merges(raw)
    else:
        raise ValueError(f"不支持的文件格式: {name}")

    df = df.replace(r"^\\N$", pd.NA, regex=True)

    return df, merges


def build_preview_html(df: pd.DataFrame, merges: list[tuple]) -> str:
    """
    根据 DataFrame 和合并区域信息，生成带 rowspan/colspan 的 HTML 预览表格。
    merges: [(min_row, min_col, max_row, max_col), ...] 0-based 数据区域坐标。
    """
    num_rows, num_cols = df.shape

    # 构建占位矩阵：记录每个格子是否被合并占据
    # occupied[(r,c)] = True 表示该位置被其他单元格的 rowspan/colspan 覆盖
    occupied = set()
    # merge_start[(r,c)] = (rowspan, colspan) 表示该位置是合并起始点
    merge_start = {}

    for min_r, min_c, max_r, max_c in merges:
        # 限制范围不超出数据
        min_r = max(0, min_r)
        min_c = max(0, min_c)
        max_r = min(max_r, num_rows - 1)
        max_c = min(max_c, num_cols - 1)
        rs = max_r - min_r + 1
        cs = max_c - min_c + 1
        if rs <= 0 or cs <= 0:
            continue
        merge_start[(min_r, min_c)] = (rs, cs)
        for dr in range(rs):
            for dc in range(cs):
                if dr == 0 and dc == 0:
                    continue
                occupied.add((min_r + dr, min_c + dc))

    # 生成 HTML
    parts = ['<table><thead><tr>']
    for col in df.columns:
        parts.append(f'<th>{_html_escape(str(col))}</th>')
    parts.append('</tr></thead><tbody>')

    for ri in range(num_rows):
        parts.append('<tr>')
        for ci in range(num_cols):
            if (ri, ci) in occupied:
                continue  # 被合并占据，跳过
            val = df.iat[ri, ci]
            text = _html_escape(_format_val(val))
            attrs = ''
            if (ri, ci) in merge_start:
                rs, cs = merge_start[(ri, ci)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            parts.append(f'<td{attrs}>{text}</td>')
        parts.append('</tr>')

    parts.append('</tbody></table>')
    return ''.join(parts)


def _html_escape(s: str) -> str:
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _format_val(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if isinstance(val, pd.Timestamp):
        return val.strftime('%Y-%m-%d %H:%M') if val.hour or val.minute else val.strftime('%Y-%m-%d')
    return str(val)


def df_to_html(df: pd.DataFrame, auto_merge: bool = True) -> str:
    """
    将 DataFrame 转为 HTML 表格字符串。
    auto_merge=True 时，自动检测左侧连续重复值的列并生成 rowspan 合并。
    合并逻辑：从左到右扫描，只有当左边所有列都相同时，当前列的相同值才合并。
    """
    num_rows, num_cols = len(df), len(df.columns)
    if num_rows == 0:
        return '<table><thead><tr>' + ''.join(f'<th>{_html_escape(str(c))}</th>' for c in df.columns) + '</tr></thead><tbody></tbody></table>'

    # 计算每个单元格的 rowspan
    # rowspan_map[(r,c)] = N 表示该位置是合并起始，跨 N 行
    # skip_set 记录被合并覆盖的位置（不输出 <td>）
    rowspan_map = {}
    skip_set = set()

    if auto_merge and num_rows > 1:
        for ci in range(num_cols):
            ri = 0
            while ri < num_rows:
                # 找连续相同值的范围
                val = _format_val(df.iat[ri, ci])
                end = ri + 1
                while end < num_rows:
                    # 检查当前值是否相同
                    next_val = _format_val(df.iat[end, ci])
                    if next_val != val or val == '':
                        break
                    # 检查左边所有列是否也在同一个合并组内
                    same_group = True
                    for left_ci in range(ci):
                        if (end, left_ci) not in skip_set and (ri, left_ci) not in rowspan_map:
                            # 左边列没有合并，当前列也不合并
                            same_group = False
                            break
                        if (end, left_ci) in skip_set:
                            # 检查是否和 ri 行属于同一个合并组
                            # 往上找合并起始行
                            found_same = False
                            for check_r in range(end - 1, -1, -1):
                                if (check_r, left_ci) in rowspan_map:
                                    rs = rowspan_map[(check_r, left_ci)]
                                    if check_r <= ri and check_r + rs > end:
                                        found_same = True
                                    break
                            if not found_same:
                                same_group = False
                                break
                    if not same_group:
                        break
                    end += 1

                span = end - ri
                if span > 1:
                    rowspan_map[(ri, ci)] = span
                    for sr in range(ri + 1, end):
                        skip_set.add((sr, ci))
                ri = end

    # 生成 HTML
    parts = ['<table><thead><tr>']
    for col in df.columns:
        parts.append(f'<th>{_html_escape(str(col))}</th>')
    parts.append('</tr></thead><tbody>')

    for ri in range(num_rows):
        parts.append('<tr>')
        for ci in range(num_cols):
            if (ri, ci) in skip_set:
                continue
            val = df.iat[ri, ci]
            text = _html_escape(_format_val(val))
            attrs = ''
            if (ri, ci) in rowspan_map:
                attrs = f' rowspan="{rowspan_map[(ri, ci)]}"'
            parts.append(f'<td{attrs}>{text}</td>')
        parts.append('</tr>')

    parts.append('</tbody></table>')
    return ''.join(parts)


def get_data_summary(df: pd.DataFrame) -> str:
    lines = [
        f"行数: {len(df)}, 列数: {len(df.columns)}",
        f"列名及类型:",
    ]
    for col in df.columns:
        dtype = df[col].dtype
        non_null = df[col].notna().sum()
        unique = df[col].nunique()
        lines.append(f"  - {col} ({dtype}): {non_null}个非空值, {unique}个唯一值")

    lines.append(f"\n前5行数据:\n{df.head().to_string()}")

    # 数值列统计
    numeric_cols = df.select_dtypes(include="number").columns
    if len(numeric_cols) > 0:
        lines.append(f"\n数值列统计:\n{df[numeric_cols].describe().to_string()}")

    return "\n".join(lines)
