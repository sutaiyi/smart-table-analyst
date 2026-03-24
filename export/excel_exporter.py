import io
from html.parser import HTMLParser

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 样式常量 ──
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="1E293B", size=11)
_SUBTOTAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
_BOLD_FONT = Font(bold=True, size=11)
_NORMAL_FONT = Font(size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1E293B")


# ── HTML 表格解析器 ──
class _TableParser(HTMLParser):
    """解析 HTML 表格，提取单元格及行级 class"""

    def __init__(self):
        super().__init__()
        self.rows = []
        self._current_row = []
        self._current_cell = None
        self._in_thead = False
        self._row_class = ""

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "thead":
            self._in_thead = True
        elif tag == "tbody":
            self._in_thead = False
        elif tag == "tr":
            self._current_row = []
            self._row_class = a.get("class", "")
        elif tag in ("td", "th"):
            self._current_cell = {
                "text": "",
                "rowspan": int(a.get("rowspan", 1)),
                "colspan": int(a.get("colspan", 1)),
                "is_header": tag == "th" or self._in_thead,
                "cell_class": a.get("class", ""),
                "row_class": self._row_class,
            }

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._current_cell is not None:
            self._current_cell["text"] = self._current_cell["text"].strip()
            self._current_row.append(self._current_cell)
            self._current_cell = None
        elif tag == "tr" and self._current_row:
            self.rows.append(self._current_row)
            self._current_row = []

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell["text"] += data


def _parse_html_table(html: str) -> list[list[dict]]:
    parser = _TableParser()
    parser.feed(html)
    return parser.rows


def _build_grid(rows: list[list[dict]]) -> list[list[dict | None]]:
    """
    将 HTML 行列表展开为完整的二维网格。
    被 rowspan/colspan 占据的格子填入 None，实际内容格子填入 cell dict。
    这是处理合并单元格的核心算法。
    """
    if not rows:
        return []

    # 第一遍：计算网格大小
    # 先用一个临时网格来确定真实列数
    num_rows = len(rows)
    # 估算最大列数（可能因 rowspan 占位而变化）
    max_cols = 0
    for row in rows:
        max_cols = max(max_cols, sum(c["colspan"] for c in row))
    # rowspan 可能扩展行数
    for row in rows:
        for cell in row:
            if cell["rowspan"] > 1:
                needed = rows.index(row) + cell["rowspan"]
                if needed > num_rows:
                    num_rows = needed

    # 构建网格，None 表示空位
    grid = [[None] * max_cols for _ in range(num_rows)]

    for row_idx, row_cells in enumerate(rows):
        col_cursor = 0
        for cell in row_cells:
            # 找到下一个空位
            while col_cursor < max_cols and grid[row_idx][col_cursor] is not None:
                col_cursor += 1
            if col_cursor >= max_cols:
                # 需要扩展列
                for g_row in grid:
                    g_row.append(None)
                max_cols += 1

            rs = cell["rowspan"]
            cs = cell["colspan"]

            # 限制 rowspan 不超出网格
            rs = min(rs, num_rows - row_idx)
            cs = min(cs, max_cols - col_cursor)

            # 在起始位置放入实际单元格
            grid[row_idx][col_cursor] = cell

            # 标记被占据的位置为 "occupied" 占位符
            for dr in range(rs):
                for dc in range(cs):
                    if dr == 0 and dc == 0:
                        continue
                    r, c = row_idx + dr, col_cursor + dc
                    if r < num_rows and c < max_cols:
                        grid[r][c] = "occupied"

            col_cursor += cs

    return grid


def _try_number(text: str):
    """尝试将文本转为数值"""
    if not text:
        return text
    cleaned = text.replace(",", "").replace(" ", "")
    # 处理百分比
    if cleaned.endswith("%"):
        try:
            return float(cleaned[:-1]) / 100
        except ValueError:
            return text
    try:
        num = float(cleaned)
        return int(num) if num == int(num) else num
    except (ValueError, OverflowError):
        return text


def _write_styled_sheet(ws, title: str, html: str):
    """将 styled_tables HTML 写入 Excel sheet，正确处理合并单元格"""
    rows = _parse_html_table(html)
    if not rows:
        return

    grid = _build_grid(rows)
    if not grid:
        return

    num_rows = len(grid)
    num_cols = len(grid[0]) if grid else 0

    # 第1行：标题
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT

    # 从第2行开始写数据
    start_row = 2

    for row_idx in range(num_rows):
        for col_idx in range(num_cols):
            cell_info = grid[row_idx][col_idx]
            excel_r = start_row + row_idx
            excel_c = col_idx + 1

            if cell_info is None:
                # 空位，写空单元格并设边框
                ws.cell(row=excel_r, column=excel_c).border = _THIN_BORDER
                continue

            if cell_info == "occupied":
                # 被合并占据，只设边框
                ws.cell(row=excel_r, column=excel_c).border = _THIN_BORDER
                continue

            # 实际内容单元格
            text = cell_info["text"]
            value = _try_number(text)
            ws_cell = ws.cell(row=excel_r, column=excel_c, value=value)
            ws_cell.border = _THIN_BORDER

            # 对齐方式
            is_num = "num" in cell_info["cell_class"] or isinstance(value, (int, float))
            h_align = "right" if is_num else "center"
            ws_cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)

            # 样式
            row_cls = cell_info["row_class"]
            rs = cell_info["rowspan"]
            cs = cell_info["colspan"]

            if cell_info["is_header"]:
                ws_cell.fill = _HEADER_FILL
                ws_cell.font = _HEADER_FONT
            elif "total" in row_cls and "subtotal" not in row_cls:
                ws_cell.fill = _TOTAL_FILL
                ws_cell.font = _BOLD_FONT
            elif "subtotal" in row_cls:
                ws_cell.fill = _SUBTOTAL_FILL
                ws_cell.font = _BOLD_FONT
            else:
                ws_cell.font = _NORMAL_FONT

            # 执行合并
            if rs > 1 or cs > 1:
                end_r = excel_r + min(rs, num_rows - row_idx) - 1
                end_c = excel_c + min(cs, num_cols - col_idx) - 1
                if end_r > excel_r or end_c > excel_c:
                    ws.merge_cells(
                        start_row=excel_r, start_column=excel_c,
                        end_row=end_r, end_column=end_c,
                    )
                    # 合并区域内所有单元格设边框
                    for mr in range(excel_r, end_r + 1):
                        for mc in range(excel_c, end_c + 1):
                            ws.cell(row=mr, column=mc).border = _THIN_BORDER

    # 自动列宽
    for col_idx in range(1, num_cols + 1):
        max_w = 8
        for row_idx in range(start_row, start_row + num_rows):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                w = len(str(val))
                if w > max_w:
                    max_w = w
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w + 4, 40)


def export_excel(
    result_tables: dict[str, pd.DataFrame],
    summary: str = "",
    styled_tables: dict[str, str] | None = None,
) -> bytes:
    buffer = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    styled = styled_tables or {}

    if summary:
        ws = wb.create_sheet("摘要")
        ws.cell(row=1, column=1, value="分析摘要").font = _TITLE_FONT
        ws.cell(row=2, column=1, value=summary).alignment = Alignment(wrap_text=True)
        ws.column_dimensions["A"].width = 80

    for title, df in result_tables.items():
        # Excel sheet名不能包含 \ / * ? : [ ] 且最长31字符
        sheet_name = title
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            sheet_name = sheet_name.replace(ch, '_')
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)

        if title in styled:
            _write_styled_sheet(ws, title, styled[title])
        else:
            # 普通 DataFrame
            ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
            for ci, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=2, column=ci, value=col_name)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
                cell.border = _THIN_BORDER
            for ri, row in enumerate(df.itertuples(index=False), 3):
                for ci, value in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=value)
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
            for ci in range(1, len(df.columns) + 1):
                max_w = len(str(df.columns[ci - 1]))
                for ri in range(3, len(df) + 3):
                    vl = len(str(ws.cell(row=ri, column=ci).value or ""))
                    if vl > max_w:
                        max_w = vl
                ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 40)

    wb.save(buffer)
    return buffer.getvalue()
